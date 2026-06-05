#!/usr/bin/env python3
# TOOL: 每日追蹤主機項目（北一區）
"""
每日追蹤主機項目 — 北一區（士林、微風、美麗華、阿波羅、高島屋、羅東）
用法：
  python3 warranty_report_n1.py              # 自動：最近完整週（週日～週六）
  python3 warranty_report_n1.py 5/31 6/4     # 指定起迄日（同年）
  python3 warranty_report_n1.py 2026-05-31 2026-06-04
"""
import subprocess, sys, os, glob, io, re, warnings
import email as _eml
from email.header import decode_header as _decode_hdr
from datetime import date, timedelta, datetime as _dt
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import JAVA, EPB_CP, EPB_CWD, MAIL_BASE, OUTPUT_BASE

# ── 門市設定 ──────────────────────────────────────────────────────────────────
STORES     = ['士林', '微風', '美麗華', '阿波羅', '高島屋', '羅東']
SHOP_CODES = {'士林':'004','微風':'005','美麗華':'024','阿波羅':'046','高島屋':'054','羅東':'057'}
SHOP_IN    = "'004','005','024','046','054','057'"

MYSETUP_PARTICIPATING = ['士林', '微風', '美麗華', '阿波羅', '高島屋']
MYSETUP_STORE_KW      = {'士林':'士林','微風':'微風','美麗華':'美麗華','阿波羅':'阿波羅','高島屋':'高島屋'}

# (key, 顯示名, header色, sub色, 目標%)
CATS = [
    ('mac',     'Mac',     '2E75B6', 'BDD7EE', 60),
    ('iphone',  'iPhone',  'C55A11', 'FCE4D6', 19),
    ('ipad',    'iPad',    '375623', 'E2EFDA', 50),
    ('watch',   'Watch',   'A50021', 'FFCCCC', 40),
    ('airpods', 'AirPods', '7030A0', 'E2CFEE', 35),
]

# ── 日期解析 ──────────────────────────────────────────────────────────────────
def parse_date(s):
    s = s.strip()
    if len(s) == 10 and s[4] == '-':
        return date.fromisoformat(s)
    if '/' in s:
        m, d = s.split('/')
        return date(date.today().year, int(m), int(d))
    raise ValueError(f"無法解析日期：{s}")

def last_saturday():
    d = date.today()
    offset = (d.weekday() + 2) % 7
    return d - timedelta(days=offset if offset else 7)

# ── EPB 查詢 ──────────────────────────────────────────────────────────────────
def epb(sql, max_rows=3000):
    r = subprocess.run(
        [JAVA, '-Dsun.net.client.defaultReadTimeout=120000',
         '-cp', EPB_CP, 'EPBReportQuery', sql, str(max_rows)],
        capture_output=True, text=True, cwd=EPB_CWD
    )
    lines = [l for l in r.stdout.strip().split('\n') if l.strip()]
    if not lines:
        return []
    hdrs = [h.strip().upper() for h in lines[0].split('\t')]
    return [dict(zip(hdrs, row.split('\t'))) for row in lines[1:]]

def query_period(d_start, d_end):
    ds   = f"TO_DATE('{d_start}','yyyy-mm-dd')"
    de   = f"TO_DATE('{d_end}','yyyy-mm-dd')"
    cond = f"l.doc_date>={ds} AND l.doc_date<={de} AND l.trans_type IN ('A','E','H')"

    def net_sql(where):
        # 排除認證機（類別2=2029）
        return (f"SELECT l.shop_id, SUM(CASE WHEN l.trans_type IN ('A','H') THEN l.stk_qty "
                f"WHEN l.trans_type='E' THEN l.stk_qty ELSE 0 END) AS units "
                f"FROM poslinev_bi l WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
                f"AND {cond} AND l.cat2_id<>'2029' AND {where} "
                f"GROUP BY l.shop_id ORDER BY l.shop_id")

    def to_map(rows):
        return {str(int(r['SHOP_ID'])): int(float(r.get('UNITS', 0) or 0)) for r in rows}

    HOST = {
        'mac':     to_map(epb(net_sql("l.cat4_id IN ('4001','4002')"))),
        'iphone':  to_map(epb(net_sql("l.cat4_id='4004'"))),
        'ipad':    to_map(epb(net_sql("l.cat4_id IN ('4005','4006','4041')"))),
        'watch':   to_map(epb(net_sql("l.cat4_id='4038'"))),
        'airpods': to_map(epb(net_sql("l.cat6_id IN ('6258','6312','6330') AND l.cat3_id='3002'"))),
    }

    C6MAP = {'6533':'mac','6534':'ipad','6535':'iphone','6536':'watch','6537':'airpods'}
    sa_rows = epb(f"SELECT l.shop_id, l.cat6_id, "
                  f"SUM(CASE WHEN l.trans_type IN ('A','H') THEN l.stk_qty "
                  f"WHEN l.trans_type='E' THEN l.stk_qty ELSE 0 END) AS units "
                  f"FROM poslinev_bi l WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
                  f"AND {cond} AND l.cat6_id IN ('6533','6534','6535','6536','6537') "
                  f"GROUP BY l.shop_id, l.cat6_id ORDER BY l.shop_id")
    sa = defaultdict(lambda: defaultdict(int))
    for row in sa_rows:
        sid = str(int(row['SHOP_ID']))
        cat = C6MAP.get(row['CAT6_ID'].strip())
        if cat:
            sa[sid][cat] += int(float(row.get('UNITS', 0) or 0))

    def acpp_cat(name):
        n = name.lower()
        if any(k in n for k in ['macbook','mac mini','imac','mac studio']): return 'mac'
        if 'iphone' in n: return 'iphone'
        if 'ipad'   in n: return 'ipad'
        if 'watch'  in n: return 'watch'
        if 'airpods'in n: return 'airpods'
        return None

    acpp_rows = epb(f"SELECT l.shop_id, l.name, "
                    f"SUM(CASE WHEN l.trans_type IN ('A','H') THEN l.stk_qty "
                    f"WHEN l.trans_type='E' THEN l.stk_qty ELSE 0 END) AS units "
                    f"FROM poslinev_bi l WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
                    f"AND {cond} AND l.cat3_id='3032' "
                    f"GROUP BY l.shop_id, l.name ORDER BY l.shop_id")
    acpp = defaultdict(lambda: defaultdict(int))
    for row in acpp_rows:
        sid = str(int(row['SHOP_ID']))
        cat = acpp_cat(row.get('NAME', ''))
        if cat:
            acpp[sid][cat] += int(float(row.get('UNITS', 0) or 0))

    return {store: {cat: (HOST[cat].get(str(int(SHOP_CODES[store])), 0),
                          sa[str(int(SHOP_CODES[store]))][cat],
                          acpp[str(int(SHOP_CODES[store]))][cat])
                    for cat, *_ in CATS}
            for store in STORES}

def query_extras(d_start, d_end):
    ds   = f"TO_DATE('{d_start}','yyyy-mm-dd')"
    de   = f"TO_DATE('{d_end}','yyyy-mm-dd')"
    cond = f"l.doc_date>={ds} AND l.doc_date<={de} AND l.trans_type IN ('A','E','H')"
    arp_rows = epb(f"SELECT l.shop_id, SUM(CASE WHEN l.trans_type IN ('A','H') THEN l.stk_qty "
                   f"WHEN l.trans_type='E' THEN l.stk_qty ELSE 0 END) AS units "
                   f"FROM poslinev_bi l WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
                   f"AND {cond} AND l.brand_id='496' GROUP BY l.shop_id ORDER BY l.shop_id")
    spk_rows = epb(f"SELECT l.shop_id, SUM(CASE WHEN l.trans_type IN ('A','H') "
                   f"THEN l.line_total_net + l.line_tax "
                   f"WHEN l.trans_type='E' THEN l.line_total_net + l.line_tax ELSE 0 END) AS amt "
                   f"FROM poslinev_bi l WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
                   f"AND {cond} AND l.cat4_id='4013' AND l.brand_id<>'453' "
                   f"GROUP BY l.shop_id ORDER BY l.shop_id")
    arp = {str(int(r['SHOP_ID'])): int(float(r.get('UNITS', 0) or 0)) for r in arp_rows}
    spk = {str(int(r['SHOP_ID'])): int(round(float(r.get('AMT', 0) or 0))) for r in spk_rows}
    return {store: {'arpedia': arp.get(str(int(SHOP_CODES[store])), 0),
                    'speaker': spk.get(str(int(SHOP_CODES[store])), 0)}
            for store in STORES}

EDU_CATS = [
    ('mac',   'Mac',   '2E75B6', 'BDD7EE'),
    ('ipad',  'iPad',  '375623', 'E2EFDA'),
    ('watch', 'Watch', 'A50021', 'FFCCCC'),
]
EDU_EXCL = {'羅東'}  # 不計算教育價佔比的門市

def query_edu_units(d_start, d_end):
    """教育價主機台數：同一單據含 SKU 99200202/99200203，且排除認證機(cat2≠2029)
    兩步驟：先取含教育SKU的單據清單，再 Python 端過濾主機台數。
    """
    ds = f"TO_DATE('{d_start}','yyyy-mm-dd')"
    de = f"TO_DATE('{d_end}','yyyy-mm-dd')"

    # Step 1: 取含教育價 SKU 的 (shop_id, doc_id) 清單
    edu_rows = epb(
        f"SELECT DISTINCT l.shop_id, l.doc_id "
        f"FROM poslinev_bi l "
        f"WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
        f"AND l.doc_date>={ds} AND l.doc_date<={de} "
        f"AND l.stk_id IN ('99200202','99200203')",
        max_rows=10000
    )
    edu_pairs = set()
    for r in edu_rows:
        try:
            edu_pairs.add((str(int(r['SHOP_ID'].strip())), r['DOC_ID'].strip()))
        except (ValueError, KeyError):
            pass

    if not edu_pairs:
        return {'mac': {}, 'ipad': {}, 'watch': {}}

    # Step 2: 查各品類的 (shop_id, doc_id, units)，Python 端過濾
    def count_edu(cat_where):
        rows = epb(
            f"SELECT l.shop_id, l.doc_id, "
            f"SUM(CASE WHEN l.trans_type IN ('A','H') THEN l.stk_qty "
            f"WHEN l.trans_type='E' THEN l.stk_qty ELSE 0 END) AS units "
            f"FROM poslinev_bi l "
            f"WHERE l.org_id='01' AND l.shop_id IN ({SHOP_IN}) "
            f"AND l.doc_date>={ds} AND l.doc_date<={de} "
            f"AND l.trans_type IN ('A','E','H') "
            f"AND l.cat2_id<>'2029' AND {cat_where} "
            f"GROUP BY l.shop_id, l.doc_id ORDER BY l.shop_id",
            max_rows=10000
        )
        result = defaultdict(int)
        for r in rows:
            try:
                sid = str(int(r['SHOP_ID'].strip()))
            except (ValueError, KeyError):
                continue
            did = r['DOC_ID'].strip()
            if (sid, did) in edu_pairs:
                result[sid] += int(float(r.get('UNITS', 0) or 0))
        return dict(result)

    return {
        'mac':   count_edu("l.cat4_id IN ('4001','4002')"),
        'ipad':  count_edu("l.cat4_id IN ('4005','4006','4041')"),
        'watch': count_edu("l.cat4_id='4038'"),
    }

# ── Mysetup 信件掃描 ──────────────────────────────────────────────────────────
def _decode_mail_str(s):
    if not s: return ''
    result = []
    for part, enc in _decode_hdr(s):
        if isinstance(part, bytes):
            result.append(part.decode(enc or 'utf-8', errors='replace'))
        else:
            result.append(part)
    return ''.join(result)

def _parse_emlx(raw):
    nl = raw.index(b'\n')
    try:
        n = int(raw[:nl].strip())
        return _eml.message_from_bytes(raw[nl + 1: nl + 1 + n])
    except Exception:
        return _eml.message_from_bytes(raw.split(b'\n', 1)[1])

def _subject_date(subject):
    m = re.search(r'(\d{1,2})-(\d{1,2})月-(\d{2})', subject)
    if not m: return None
    try:
        return date(2000 + int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None

def query_mysetup(d_start, d_end):
    emlx_files = glob.glob(os.path.join(str(MAIL_BASE), '**/Messages/*.emlx'), recursive=True)
    emlx_files = [f for f in emlx_files if not f.endswith('.partial.emlx')]
    lo = _dt(d_start.year, d_start.month, d_start.day).timestamp() - 86400
    hi = _dt(d_end.year, d_end.month, d_end.day).timestamp() + 86400 * 2
    emlx_files = [f for f in emlx_files if lo <= os.path.getmtime(f) <= hi]

    counts = defaultdict(int)
    seen_dates = set()
    for path in emlx_files:
        try:
            with open(path, 'rb') as f:
                raw = f.read()
            if b'\n' not in raw: continue
            msg = _parse_emlx(raw)
            subj = _decode_mail_str(msg.get('Subject', ''))
            if 'Personal Setup' not in subj or 'Setup Data' not in subj: continue
            data_date = _subject_date(subj)
            if data_date is None or not (d_start <= data_date <= d_end): continue
            if data_date in seen_dates: continue
            seen_dates.add(data_date)
            for part in msg.walk():
                fn = _decode_mail_str(part.get_filename() or '')
                if not fn.lower().endswith('.xlsx'): continue
                raw_xls = part.get_payload(decode=True)
                if not raw_xls: continue
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    wb = load_workbook(io.BytesIO(raw_xls), read_only=True, data_only=True)
                for sheet in wb.worksheets:
                    headers = None
                    pos_col = None
                    for row in sheet.iter_rows(values_only=True):
                        if headers is None:
                            headers = [str(v).strip() if v else '' for v in row]
                            if 'POS Name' in headers:
                                pos_col = headers.index('POS Name')
                            else:
                                break
                            continue
                        if pos_col is None: break
                        pos = str(row[pos_col]) if row[pos_col] else ''
                        for store, kw in MYSETUP_STORE_KW.items():
                            if kw in pos:
                                counts[store] += 1
                                break
                wb.close()
                break
        except Exception:
            continue
    return dict(counts)

# ── Excel 輸出 ────────────────────────────────────────────────────────────────
thin = Side(style='thin', color='BFBFBF')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(ws, r, c, val=None, bg=None, fg='000000', bold=False, sz=10):
    cell = ws.cell(row=r, column=c)
    cell.value = val
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORD
    return cell

def write_table(ws, start_row, title, data, extras=None, mysetup=None):
    mysetup_col = 2 + len(CATS) * 4 + (2 if extras else 0)
    total_cols  = 1 + len(CATS) * 4 + (2 if extras else 0) + (2 if mysetup is not None else 0)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    sc(ws, start_row, 1, title, bg='1F3864', fg='FFFFFF', bold=True, sz=12)
    ws.row_dimensions[start_row].height = 40

    r2 = start_row + 1
    sc(ws, r2, 1, '門市', bg='404040', fg='FFFFFF', bold=True)
    ws.row_dimensions[r2].height = 22
    col = 2
    for cat, label, hbg, sbg, tgt in CATS:
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 3)
        sc(ws, r2, col, f'{label}  (目標 {tgt}%)', bg=hbg, fg='FFFFFF', bold=True, sz=10)
        col += 4
    if extras:
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 1)
        sc(ws, r2, col, 'ARpedia / 喇叭', bg='404040', fg='FFFFFF', bold=True, sz=10)
    if mysetup is not None:
        ws.merge_cells(start_row=r2, start_column=mysetup_col, end_row=r2, end_column=mysetup_col + 1)
        sc(ws, r2, mysetup_col, 'Mysetup 提交率', bg='1F4E79', fg='FFFFFF', bold=True, sz=10)

    r3 = start_row + 2
    sc(ws, r3, 1, '', bg='595959')
    ws.row_dimensions[r3].height = 32
    col = 2
    for cat, label, hbg, sbg, tgt in CATS:
        for h in ['台數', 'ACPP+', 'SACare', '搭售率']:
            sc(ws, r3, col, h, bg=sbg, bold=True, sz=9)
            col += 1
    if extras:
        sc(ws, r3, col,     'ARpedia\n數量', bg='D9D9D9', bold=True, sz=9)
        sc(ws, r3, col + 1, '喇叭\n金額',    bg='D9D9D9', bold=True, sz=9)
    if mysetup is not None:
        sc(ws, r3, mysetup_col,     'Mysetup\n提交數', bg='DEEAF1', bold=True, sz=9)
        sc(ws, r3, mysetup_col + 1, '提交率',          bg='DEEAF1', bold=True, sz=9)

    for i, store in enumerate(STORES):
        r  = start_row + 3 + i
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F5F5'
        ws.row_dimensions[r].height = 20
        sc(ws, r, 1, store, bg=bg, bold=True)
        col = 2
        for cat, label, hbg, sbg, tgt in CATS:
            host, sa_, acpp_ = data[store][cat]
            w   = sa_ + acpp_
            pct = w / host * 100 if host > 0 else None
            ps  = f'{pct:.0f}%' if pct is not None else '—'
            txt = ('C00000' if pct is not None and pct < tgt else '375623' if pct is not None else 'AAAAAA')
            sc(ws, r, col,     host if host > 0 else 0,     bg=bg)
            sc(ws, r, col + 1, acpp_ if acpp_ != 0 else '', bg=bg)
            sc(ws, r, col + 2, sa_   if sa_   != 0 else '', bg=bg)
            sc(ws, r, col + 3, ps, bg=bg, fg=txt, bold=True)
            col += 4
        if extras:
            sc(ws, r, col,     extras[store]['arpedia'] or '', bg=bg)
            sc(ws, r, col + 1, extras[store]['speaker']  or '', bg=bg)
        if mysetup is not None:
            if store in MYSETUP_PARTICIPATING:
                ms    = mysetup.get(store, 0)
                denom = sum(data[store][c][0] for c, *_ in CATS if c != 'airpods')
                pv    = ms / denom * 100 if denom > 0 else None
                ps    = f'{pv:.0f}%' if pv is not None else '—'
                txt   = ('C00000' if pv < 60 else '375623') if pv is not None else 'AAAAAA'
                sc(ws, r, mysetup_col,     ms, bg=bg)
                sc(ws, r, mysetup_col + 1, ps, bg=bg, fg=txt, bold=True)
            else:
                sc(ws, r, mysetup_col,     '—', bg=bg, fg='AAAAAA')
                sc(ws, r, mysetup_col + 1, '—', bg=bg, fg='AAAAAA')

    r_tot = start_row + 3 + len(STORES)
    ws.row_dimensions[r_tot].height = 22
    sc(ws, r_tot, 1, '合計', bg='404040', fg='FFFFFF', bold=True)
    col = 2
    for cat, label, hbg, sbg, tgt in CATS:
        t_h = sum(data[s][cat][0] for s in STORES)
        t_w = sum(data[s][cat][1] + data[s][cat][2] for s in STORES)
        pct = t_w / t_h * 100 if t_h > 0 else None
        ps  = f'{pct:.0f}%' if pct is not None else '—'
        txt = ('C00000' if pct is not None and pct < tgt else '375623' if pct is not None else 'AAAAAA')
        sc(ws, r_tot, col,     t_h, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 1, sum(data[s][cat][2] for s in STORES) or '', bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 2, sum(data[s][cat][1] for s in STORES), bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 3, ps, fg=txt, bg='D9D9D9', bold=True)
        col += 4
    if extras:
        sc(ws, r_tot, col,     sum(extras[s]['arpedia'] for s in STORES), bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 1, sum(extras[s]['speaker']  for s in STORES), bg='D9D9D9', bold=True)
    if mysetup is not None:
        t_ms    = sum(mysetup.get(s, 0) for s in MYSETUP_PARTICIPATING)
        t_denom = sum(data[s][c][0] for s in MYSETUP_PARTICIPATING for c, *_ in CATS if c != 'airpods')
        pv      = t_ms / t_denom * 100 if t_denom > 0 else None
        ps      = f'{pv:.0f}%' if pv is not None else '—'
        txt     = ('C00000' if pv < 60 else '375623') if pv is not None else 'AAAAAA'
        sc(ws, r_tot, mysetup_col,     t_ms, bg='D9D9D9', bold=True)
        sc(ws, r_tot, mysetup_col + 1, ps,   bg='D9D9D9', bold=True, fg=txt)

    return r_tot + 2

def write_edu_table(ws, start_row, title, edu_data, host_data):
    total_cols = 1 + len(EDU_CATS) * 3

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    sc(ws, start_row, 1, title, bg='1F3864', fg='FFFFFF', bold=True, sz=12)
    ws.row_dimensions[start_row].height = 28

    r2 = start_row + 1
    sc(ws, r2, 1, '門市', bg='404040', fg='FFFFFF', bold=True)
    ws.row_dimensions[r2].height = 22
    col = 2
    for cat, label, hbg, sbg in EDU_CATS:
        ws.merge_cells(start_row=r2, start_column=col, end_row=r2, end_column=col + 2)
        sc(ws, r2, col, label, bg=hbg, fg='FFFFFF', bold=True)
        col += 3

    r3 = start_row + 2
    sc(ws, r3, 1, '', bg='595959')
    ws.row_dimensions[r3].height = 28
    col = 2
    for cat, label, hbg, sbg in EDU_CATS:
        sc(ws, r3, col,     '教育台數', bg=sbg, bold=True, sz=9)
        sc(ws, r3, col + 1, '總台數',   bg=sbg, bold=True, sz=9)
        sc(ws, r3, col + 2, '教育佔比', bg=sbg, bold=True, sz=9)
        col += 3

    for i, store in enumerate(STORES):
        r  = start_row + 3 + i
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F5F5'
        ws.row_dimensions[r].height = 20
        sc(ws, r, 1, store, bg=bg, bold=True)
        col = 2
        if store in EDU_EXCL:
            for _ in EDU_CATS:
                sc(ws, r, col,     '—', bg=bg, fg='AAAAAA')
                sc(ws, r, col + 1, '—', bg=bg, fg='AAAAAA')
                sc(ws, r, col + 2, '—', bg=bg, fg='AAAAAA')
                col += 3
        else:
            sid = str(int(SHOP_CODES[store]))
            for cat, *_ in EDU_CATS:
                edu_u = edu_data[cat].get(sid, 0)
                tot_u = host_data[store][cat][0]
                pct   = edu_u / tot_u * 100 if tot_u > 0 else None
                ps    = f'{pct:.0f}%' if pct is not None else '—'
                sc(ws, r, col,     edu_u if edu_u > 0 else 0, bg=bg)
                sc(ws, r, col + 1, tot_u if tot_u > 0 else 0, bg=bg)
                sc(ws, r, col + 2, ps, bg=bg, bold=True)
                col += 3

    incl = [s for s in STORES if s not in EDU_EXCL]
    r_tot = start_row + 3 + len(STORES)
    ws.row_dimensions[r_tot].height = 22
    sc(ws, r_tot, 1, '合計', bg='404040', fg='FFFFFF', bold=True)
    col = 2
    for cat, *_ in EDU_CATS:
        t_edu = sum(edu_data[cat].get(str(int(SHOP_CODES[s])), 0) for s in incl)
        t_tot = sum(host_data[s][cat][0] for s in incl)
        pct   = t_edu / t_tot * 100 if t_tot > 0 else None
        ps    = f'{pct:.0f}%' if pct is not None else '—'
        sc(ws, r_tot, col,     t_edu, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 1, t_tot, bg='D9D9D9', bold=True)
        sc(ws, r_tot, col + 2, ps,    bg='D9D9D9', bold=True)
        col += 3

    return r_tot + 2


def set_col_widths(ws):
    ws.column_dimensions['A'].width = 7
    col = 2
    for _ in CATS:
        for w in [6, 7, 7, 7]:
            ws.column_dimensions[get_column_letter(col)].width = w
            col += 1
    ws.column_dimensions[get_column_letter(col)].width     = 9
    ws.column_dimensions[get_column_letter(col + 1)].width = 11
    ws.column_dimensions[get_column_letter(col + 2)].width = 8
    ws.column_dimensions[get_column_letter(col + 3)].width = 7

# ── 主程式 ────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) >= 3:
        d_start, d_end = parse_date(sys.argv[1]), parse_date(sys.argv[2])
    elif len(sys.argv) == 2:
        d_end   = parse_date(sys.argv[1])
        d_start = d_end.replace(day=1)
    else:
        sat     = last_saturday()
        d_start = sat - timedelta(days=6)
        d_end   = sat

    m_start = d_end.replace(day=1)
    print(f"區間：{d_start.strftime('%-m/%-d')} ～ {d_end.strftime('%-m/%-d')}")
    print(f"月累積：{m_start.strftime('%-m/%-d')} ～ {d_end.strftime('%-m/%-d')}")

    print("查詢區間資料... (1/5)", flush=True)
    data_range = query_period(d_start, d_end)
    print("查詢月累積資料... (2/5)", flush=True)
    data_month = query_period(m_start, d_end)
    print("查詢 ARpedia/喇叭資料... (3/5)", flush=True)
    extras_range = query_extras(d_start, d_end)
    extras_month = query_extras(m_start, d_end)
    print("掃描 Mysetup 信件... (4/5)", flush=True)
    mysetup_range = query_mysetup(d_start, d_end)
    mysetup_month = query_mysetup(m_start, d_end)
    print(f"  找到門市：{', '.join(sorted(mysetup_range)) or '（無）'}")
    print("查詢教育價台數... (5/5)", flush=True)
    edu_range = query_edu_units(d_start, d_end)
    edu_month = query_edu_units(m_start, d_end)

    wb = Workbook()
    ws = wb.active
    ws.title = '每日追蹤主機項目'
    ws.freeze_panes = 'B4'

    NOTE = '※ 主機台數已排除認證機（類別2=2029）'
    rt = (f"區間  {d_start.strftime('%-m/%-d')} ～ {d_end.strftime('%-m/%-d')}  "
          f"各門市保固搭售率（北一區）\n{NOTE}")
    mt = (f"月累積  {m_start.strftime('%-m/%d').lstrip('0')} ～ {d_end.strftime('%-m/%-d')}  "
          f"各門市保固搭售率（北一區）\n{NOTE}")
    next_row = write_table(ws, 1, rt, data_range, extras_range, mysetup_range)
    write_table(ws, next_row, mt, data_month, extras_month, mysetup_month)
    set_col_widths(ws)

    ws2 = wb.create_sheet('教育價佔比')
    ws2.freeze_panes = 'B4'
    rt2 = f"區間  {d_start.strftime('%-m/%-d')} ～ {d_end.strftime('%-m/%-d')}  教育價佔比（北一區）"
    mt2 = f"月累積  {m_start.strftime('%-m/%d').lstrip('0')} ～ {d_end.strftime('%-m/%-d')}  教育價佔比（北一區）"
    next_row2 = write_edu_table(ws2, 1, rt2, edu_range, data_range)
    write_edu_table(ws2, next_row2, mt2, edu_month, data_month)
    ws2.column_dimensions['A'].width = 9
    for i in range(len(EDU_CATS) * 3):
        ws2.column_dimensions[get_column_letter(2 + i)].width = 9

    out_dir = OUTPUT_BASE / '北一區'
    out_dir.mkdir(parents=True, exist_ok=True)
    fname   = f"保固搭售率_{d_start.strftime('%m%d')}-{d_end.strftime('%m%d')}.xlsx"
    out_path = out_dir / fname
    wb.save(out_path)
    print(f"\n✅ 完成：{out_path}")
    subprocess.run(['open', str(out_path)])

if __name__ == '__main__':
    main()
