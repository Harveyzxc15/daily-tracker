#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PQI點數兌換追蹤 — 北一＋北二區

來源：EPB 140EB / STOREDTL
  ‧ 兌換數量 = src_code='RPOSN' 的出庫量（指定日期區間，依 loc_id 彙總）
  ‧ 目前在庫 = STOREDTL 全期淨異動，依實體倉 store_id（SA+門市碼）彙總

輸出：同一工作表上下兩塊，左為存貨代碼+品名(含點數)，上為門市。
需 VPN + 已安裝 EPBrowser。

★ 日期區間請改下方 START_DATE / END_DATE 兩個變數。
"""
import sys
import subprocess
from datetime import date

from config import JAVA, EPB_CP, EPB_CWD, OUTPUT_BASE

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
#  ★★★ 兌換查詢的日期區間 ★★★
#  ‧ 從選單列點「指定區間…」會用對話框輸入，蓋過下面這兩個。
#  ‧ 直接跑腳本（沒帶參數）時，就用下面這兩個。
START_DATE = "2026-06-01"   # 起（含）
END_DATE   = "2026-06-20"   # 迄（含）
# ═══════════════════════════════════════════════════════════════════


def _norm_date(s):
    """把 'M/D'、'YYYY/M/D'、'YYYY-M-D' 統一成 'YYYY-MM-DD'。"""
    s = s.strip().replace("/", "-")
    parts = s.split("-")
    if len(parts) == 2:                       # M-D → 補今年
        parts = [str(date.today().year)] + parts
    y, m, d = (int(p) for p in parts)
    return f"{y:04d}-{m:02d}-{d:02d}"


# 命令列參數可覆蓋日期：python3 PQI點數兌換追蹤.py <起> <迄>
if len(sys.argv) >= 3:
    START_DATE = _norm_date(sys.argv[1])
    END_DATE   = _norm_date(sys.argv[2])
elif len(sys.argv) == 2:                       # 只給一個 → 起迄同一天
    START_DATE = END_DATE = _norm_date(sys.argv[1])

# 追蹤品項：(存貨代碼, 品名, 點數)
ITEMS = [
    ("27700462", "@PQI 6000E 輕巧行動電源-白色",            99),
    ("27700562", "PQI 18W三合一無線立充",                  199),
    ("27700648", "PQI WCS23WR 23W三合一磁吸無線充電座",     250),
    ("27700651", "PQI WCS15W 15W快充磁吸折疊充電座",         99),
    ("27700639", "PQI PD10 10000雙向快充行動電源",          299),
    ("27700668", "PQI WCC2301 MagSafe 三合一摺疊充電座",    250),
]

# 門市：(門市碼, 店名)；北一前 N1_COUNT 間、其餘北二
STORES = [
    ("004", "士林"), ("005", "微風"), ("024", "美麗華"),
    ("046", "阿波羅"), ("054", "高島屋"), ("057", "羅東"),
    ("009", "永和"), ("025", "板橋誠品"), ("050", "新西門"),
    ("055", "花蓮"), ("063", "板橋遠百"), ("064", "新莊宏匯"),
    ("068", "新店裕隆城"),
]
N1_COUNT = 6

# 個人清單要排除的非真人帳號（SA999=總公司）
EXCLUDE_EMPS = ["SA999"]


def run_sql(sql, max_rows=2000):
    """執行 EPB 查詢，回傳 list[dict]（首列為欄名）。"""
    r = subprocess.run([JAVA, "-cp", EPB_CP, "EPBReportQuery", sql, str(max_rows)],
                       capture_output=True, text=True, cwd=EPB_CWD)
    lines = [l for l in r.stdout.strip().split("\n") if l]
    if not lines:
        if r.stderr.strip():
            raise SystemExit("EPB 查詢失敗（VPN 沒連？）：\n" + r.stderr.strip()[:600])
        return []
    hdr = lines[0].split("\t")
    return [dict(zip(hdr, ln.split("\t"))) for ln in lines[1:]]


def query_data():
    codes = ",".join(f"'{c}' " for c, _, _ in ITEMS)
    locs = ",".join(f"'{c}'" for c, _ in STORES)
    sa = ",".join(f"'SA{c}'" for c, _ in STORES)

    # 兌換（RPOSN 出庫，stk_qty 為負 → 取絕對值），依 loc_id
    redeem_rows = run_sql(
        f"SELECT loc_id, stk_id, SUM(stk_qty) qty FROM storedtl "
        f"WHERE src_code='RPOSN' AND stk_id IN ({codes}) "
        f"AND doc_date >= TO_DATE('{START_DATE}','YYYY-MM-DD') "
        f"AND doc_date <= TO_DATE('{END_DATE}','YYYY-MM-DD') "
        f"AND loc_id IN ({locs}) GROUP BY loc_id, stk_id")

    # 在庫（全期淨異動），依實體倉 store_id = SA+門市碼
    stock_rows = run_sql(
        f"SELECT store_id, stk_id, SUM(stk_qty) qty FROM storedtl "
        f"WHERE stk_id IN ({codes}) AND store_id IN ({sa}) "
        f"GROUP BY store_id, stk_id")

    redeem = {stk: {} for stk, _, _ in ITEMS}
    for r in redeem_rows:
        redeem[r["STK_ID"]][r["LOC_ID"]] = abs(int(float(r["QTY"])))

    stock = {stk: {} for stk, _, _ in ITEMS}
    for r in stock_rows:
        code = r["STORE_ID"][2:]  # 去掉 'SA'
        stock[r["STK_ID"]][code] = int(float(r["QTY"]))
    return redeem, stock


def query_emp():
    """個人兌換明細：回傳 [(門市碼, 員工代碼, 姓名, {stk_id: 數量}), ...]

    名冊＝當期在該店有「銷售(POSN)」或「兌換(RPOSN)」紀錄的人員之聯集，
    因此沒兌換的人也會在清單裡（各品項空白、合計 0）。
    """
    codes = ",".join(f"'{c}'" for c, _, _ in ITEMS)
    locs = ",".join(f"'{c}'" for c, _ in STORES)
    period = (f"d.doc_date >= TO_DATE('{START_DATE}','YYYY-MM-DD') "
              f"AND d.doc_date <= TO_DATE('{END_DATE}','YYYY-MM-DD')")
    excl = ",".join(f"'{e}'" for e in EXCLUDE_EMPS)

    # 名冊：當期有銷售或兌換紀錄的門市人員
    roster = run_sql(
        f"SELECT DISTINCT d.loc_id, d.emp_id, "
        f"(SELECT MAX(e.name) FROM ep_emp e WHERE e.emp_id=d.emp_id) nm "
        f"FROM storedtl d WHERE d.src_code IN ('POSN','RPOSN') AND {period} "
        f"AND d.loc_id IN ({locs}) AND d.emp_id IS NOT NULL "
        f"AND d.emp_id NOT IN ({excl})")

    # 兌換明細
    rows = run_sql(
        f"SELECT d.loc_id, d.emp_id, d.stk_id, SUM(d.stk_qty) qty FROM storedtl d "
        f"WHERE d.src_code='RPOSN' AND d.stk_id IN ({codes}) AND {period} "
        f"AND d.loc_id IN ({locs}) GROUP BY d.loc_id, d.emp_id, d.stk_id")

    people = {(r["LOC_ID"], r["EMP_ID"], r.get("NM") or ""): {} for r in roster}
    names = {(r["LOC_ID"], r["EMP_ID"]): r.get("NM") or "" for r in roster}
    for r in rows:
        key = (r["LOC_ID"], r["EMP_ID"], names.get((r["LOC_ID"], r["EMP_ID"]), ""))
        people.setdefault(key, {})[r["STK_ID"]] = abs(int(float(r["QTY"])))

    order = {c: i for i, (c, _) in enumerate(STORES)}
    out = [(loc, emp, nm, d) for (loc, emp, nm), d in people.items()]
    # 依門市順序排，店內再依兌換總數由多到少（未兌換者自然排在該店最後）
    out.sort(key=lambda x: (order.get(x[0], 99), -sum(x[3].values()), x[1]))
    return out


# ── Excel 樣式 ──────────────────────────────────────────────────────
HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")
N1FILL = PatternFill("solid", fgColor="DDEBF7")
N2FILL = PatternFill("solid", fgColor="FCE4D6")
ZEROFILL = PatternFill("solid", fgColor="FFC7CE")   # 未兌換人員整列標色
TITLE = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center", vertical="center")
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(_THIN, _THIN, _THIN, _THIN)
TOTAL_COL = 3 + len(STORES)


def write_block(ws, title, dataset, start):
    ws.cell(start, 1, title).font = TITLE
    h1, h2, first = start + 1, start + 2, start + 3
    ws.cell(h1, 1, "存貨代碼").font = HDR
    ws.cell(h1, 2, "品名").font = HDR
    ws.cell(h1, 1).fill = HFILL
    ws.cell(h1, 2).fill = HFILL
    ws.cell(h1, 3, "北一區").font = Font(bold=True)
    ws.cell(h1, 3).fill = N1FILL
    ws.cell(h1, 3).alignment = CENTER
    ws.merge_cells(start_row=h1, start_column=3, end_row=h1, end_column=2 + N1_COUNT)
    c1 = 3 + N1_COUNT
    ws.cell(h1, c1, "北二區").font = Font(bold=True)
    ws.cell(h1, c1).fill = N2FILL
    ws.cell(h1, c1).alignment = CENTER
    ws.merge_cells(start_row=h1, start_column=c1, end_row=h1, end_column=2 + len(STORES))
    ws.cell(h1, TOTAL_COL, "合計").font = HDR
    ws.cell(h1, TOTAL_COL).fill = HFILL

    ws.cell(h2, 1, "").fill = HFILL
    ws.cell(h2, 2, "").fill = HFILL
    for i, (code, name) in enumerate(STORES):
        cc = ws.cell(h2, 3 + i, name)
        cc.font = Font(bold=True)
        cc.alignment = CENTER
        cc.fill = N1FILL if i < N1_COUNT else N2FILL
    ws.cell(h2, TOTAL_COL, "").fill = HFILL

    for r, (stk, nm, pt) in enumerate(ITEMS):
        row = first + r
        ws.cell(row, 1, stk)
        ws.cell(row, 2, f"{nm} {pt}點")
        for i, (code, _) in enumerate(STORES):
            v = dataset[stk].get(code)
            ws.cell(row, 3 + i, v if v is not None else None).alignment = CENTER
        f, l = get_column_letter(3), get_column_letter(2 + len(STORES))
        ws.cell(row, TOTAL_COL, f"=SUM({f}{row}:{l}{row})").alignment = CENTER

    trow = first + len(ITEMS)
    ws.cell(trow, 1, "合計").font = Font(bold=True)
    for i in range(len(STORES)):
        col = get_column_letter(3 + i)
        ws.cell(trow, 3 + i, f"=SUM({col}{first}:{col}{first + len(ITEMS) - 1})").font = Font(bold=True)
        ws.cell(trow, 3 + i).alignment = CENTER
    tc = get_column_letter(TOTAL_COL)
    ws.cell(trow, TOTAL_COL, f"=SUM({tc}{first}:{tc}{first + len(ITEMS) - 1})").font = Font(bold=True)
    ws.cell(trow, TOTAL_COL).alignment = CENTER

    for rr in range(h1, trow + 1):
        for cc in range(1, TOTAL_COL + 1):
            ws.cell(rr, cc).border = BORDER
    return trow


def write_emp_sheet(ws, people, region, fill):
    """個人兌換清單（單一區）：一列一位人員，欄為各品項。"""
    names = {c: n for c, n in STORES}
    nzero = sum(1 for p in people if sum(p[3].values()) == 0)
    ws.cell(1, 1, f"{region}人員兌換清單（RPOSN，{START_DATE} ~ {END_DATE}）　"
                  f"共 {len(people)} 人，其中 {nzero} 人未兌換（紅底）").font = TITLE
    h = 2
    for i, t in enumerate(["門市", "員工代碼", "姓名"]):
        c = ws.cell(h, 1 + i, t)
        c.font = HDR
        c.fill = HFILL
        c.alignment = CENTER
    for i, (_, nm, pt) in enumerate(ITEMS):
        c = ws.cell(h, 4 + i, f"{nm}\n{pt}點")
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = fill
    tcol = 4 + len(ITEMS)
    ws.cell(h, tcol, "合計").font = HDR
    ws.cell(h, tcol).fill = HFILL
    ws.cell(h, tcol).alignment = CENTER
    ws.row_dimensions[h].height = 58

    first = h + 1
    for r, (loc, emp, nm, d) in enumerate(people):
        row = first + r
        zero = sum(d.values()) == 0          # 當期完全沒兌換
        ws.cell(row, 1, names.get(loc, loc)).alignment = CENTER
        ws.cell(row, 2, emp).alignment = CENTER
        ws.cell(row, 3, nm)
        ws.cell(row, 1).fill = fill
        for i, (stk, _, _) in enumerate(ITEMS):
            v = d.get(stk)
            ws.cell(row, 4 + i, v if v else None).alignment = CENTER
        f, l = get_column_letter(4), get_column_letter(3 + len(ITEMS))
        ws.cell(row, tcol, f"=SUM({f}{row}:{l}{row})").alignment = CENTER
        ws.cell(row, tcol).font = Font(bold=True)
        if zero:                              # 未兌換整列標紅底
            for cc in range(1, tcol + 1):
                ws.cell(row, cc).fill = ZEROFILL

    trow = first + len(people)
    ws.cell(trow, 1, "合計").font = Font(bold=True)
    for i in range(len(ITEMS) + 1):
        col = get_column_letter(4 + i)
        ws.cell(trow, 4 + i, f"=SUM({col}{first}:{col}{trow - 1})").font = Font(bold=True)
        ws.cell(trow, 4 + i).alignment = CENTER

    for rr in range(h, trow + 1):
        for cc in range(1, tcol + 1):
            ws.cell(rr, cc).border = BORDER
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for i in range(len(ITEMS)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 11
    ws.column_dimensions[get_column_letter(tcol)].width = 8
    ws.freeze_panes = "D3"


def build_excel(redeem, stock, people):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PQI點數兌換追蹤"
    end1 = write_block(ws, f"兌換數量（來源代碼 RPOSN，{START_DATE} ~ {END_DATE}）", redeem, 1)
    write_block(ws, f"目前在庫數量（140EB STOREDTL 淨異動，截至 {date.today():%Y/%m/%d}）", stock, end1 + 2)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    for i in range(len(STORES)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 8
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 8
    ws.freeze_panes = "C3"

    n1_locs = {c for c, _ in STORES[:N1_COUNT]}
    write_emp_sheet(wb.create_sheet("個人兌換清單-北一"),
                    [p for p in people if p[0] in n1_locs], "北一區", N1FILL)
    write_emp_sheet(wb.create_sheet("個人兌換清單-北二"),
                    [p for p in people if p[0] not in n1_locs], "北二區", N2FILL)

    OUTPUT_BASE.mkdir(parents=True, exist_ok=True)
    out = OUTPUT_BASE / f"PQI點數兌換追蹤_{START_DATE.replace('-','')}-{END_DATE.replace('-','')}.xlsx"
    wb.save(out)
    return out


def main():
    print(f"查詢 {START_DATE} ~ {END_DATE} …")
    redeem, stock = query_data()
    people = query_emp()
    print(f"人員 {len(people)} 位")
    out = build_excel(redeem, stock, people)
    print("已輸出：", out)
    subprocess.run(["open", str(out)])


if __name__ == "__main__":
    main()
